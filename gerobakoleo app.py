import streamlit as st
import json
import os
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import openpyxl
import glob 

# ================= 1. KONFIGURASI UTAMA =================
TOKEN_FONNTE    = "VP1u4odNqETyKTs8mXp4"  # Token Fonnte Kamu
TARGET_WA       = "120363406910541987@g.us" # Ganti dengan ID Grup WA
PIN_OWNER_LOGIN = "8888" # PIN untuk Login ke Menu Owner
PASSWORD_RESET  = "ciroclistopel" # Password khusus untuk Hapus Data

# ================= 2. DATABASE & FILE =================
FILE_DB_GEROBAK     = "database_gerobak.json" 
FILE_DB_STAFF       = "database_staff.json"   
FILE_DB_MENU        = "database_menu.json"    
FILE_DB_LOKASI      = "database_lokasi.json"  
FILE_DB_SURAT_JALAN = "database_surat_jalan.json"

# --- DATA DEFAULT (3 KATEGORI) ---
MENU_DEFAULT = {
    "Regular (Cup)": {
        "Fresh Milk": 8000, 
        "Coklat Milk": 10000,
        "Strawberry Milk": 10000,
        "Vanilla Milk": 10000,
        "Mango Milk": 10000,
        "Melon Milk": 10000
    },
    "Botol 250ml": {
        "Fresh Milk": 10000, 
        "Coklat Milk": 15000,
        "Strawberry Milk": 15000,
        "Vanilla Milk": 15000,
        "Mango Milk": 15000,
        "Melon Milk": 15000
    },
    "Botol 1 Liter": {
        "Plastic Edition": 20000, 
        "Coklat Milk": 45000,
        "Strawberry Milk": 45000,
        "Vanilla Milk": 45000,
        "Mango Milk": 45000,
        "Melon Milk": 45000
    }
}

LOKASI_DEFAULT = {
    "1": "Gerobak 01 - SD Kartika", 
    "2": "Gerobak 02 - belum ada", 
    "3": "Gerobak 03 - belum ada"
}

# ================= 3. FUNGSI BANTUAN =================
def get_wib_now():
    return datetime.utcnow() + timedelta(hours=7)

def format_rupiah(angka):
    return f"Rp {int(angka):,}".replace(",", ".")

# --- FUNGSI BARU: KIRIM KE WHATSAPP VIA FONNTE ---
def kirim_whatsapp(pesan):
    if "PASTE_ID" in TARGET_WA: return # Cegah error jika ID belum diisi
    try:
        url = "https://api.fonnte.com/send"
        headers = {
            'Authorization': TOKEN_FONNTE
        }
        data = {
            'target': TARGET_WA,
            'message': pesan
        }
        requests.post(url, headers=headers, data=data, timeout=5)
    except Exception as e:
        print(f"Gagal kirim WA: {e}")

def load_json(filename):
    if os.path.exists(filename):
        try:
            with open(filename, 'r') as f: return json.load(f)
        except json.JSONDecodeError: return {}
    return {}

def save_json(filename, data):
    with open(filename, 'w') as f: json.dump(data, f, indent=4)

# --- FUNGSI UPDATE DATA ---
def get_lokasi_aktif():
    data = load_json(FILE_DB_LOKASI)
    if not data: save_json(FILE_DB_LOKASI, LOKASI_DEFAULT); return LOKASI_DEFAULT
    return data

def simpan_lokasi_baru(id_lokasi, nama_lengkap_lokasi):
    data = get_lokasi_aktif(); data[str(id_lokasi)] = nama_lengkap_lokasi; save_json(FILE_DB_LOKASI, data)

def hapus_lokasi(id_lokasi):
    data = get_lokasi_aktif()
    if str(id_lokasi) in data: del data[str(id_lokasi)]; save_json(FILE_DB_LOKASI, data)

# --- FUNGSI MENU BARU (BERTINGKAT) ---
def get_menu_aktif():
    data = load_json(FILE_DB_MENU)
    if not data or not isinstance(list(data.values())[0], dict): 
        save_json(FILE_DB_MENU, MENU_DEFAULT)
        return MENU_DEFAULT
    return data

def simpan_menu_baru(kategori, nama, harga):
    data = get_menu_aktif()
    if kategori not in data: data[kategori] = {}
    data[kategori][nama] = int(harga)
    save_json(FILE_DB_MENU, data)

def hapus_menu(kategori, nama):
    data = get_menu_aktif()
    if kategori in data and nama in data[kategori]:
        del data[kategori][nama]
        save_json(FILE_DB_MENU, data)

def simpan_staff_baru(nama, pin):
    data = load_json(FILE_DB_STAFF); 
    if pin in data: return False
    data[pin] = nama; save_json(FILE_DB_STAFF, data); return True

def hapus_staff(pin):
    data = load_json(FILE_DB_STAFF)
    if pin in data: del data[pin]; save_json(FILE_DB_STAFF, data); return True
    return False

# ================= 4. FUNGSI EXCEL =================
def get_nama_file_excel(nama_staff):
    nama_clean = nama_staff.replace(" ", "_").upper()
    return f"LAPORAN_{nama_clean}.xlsx"

def rapikan_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws[1]:
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = thin_border

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            header_cell = ws[f"{col_letter}1"]
            header_text = str(header_cell.value).upper() if header_cell.value else ""
            
            is_currency = any(x in header_text for x in ['CASH', 'QRIS', 'TOTAL'])
            is_menu_col = "(" in header_text and ")" in header_text and "-" in header_text

            for cell in col:
                cell.border = thin_border
                if is_currency and cell.row > 1:
                    try: 
                        if isinstance(cell.value, (int, float)): cell.number_format = '"Rp" #,##0'
                    except: pass
                
                if is_menu_col and cell.row > 1: cell.alignment = center

                try:
                    if cell.value and len(str(cell.value)) > max_len: max_len = len(str(cell.value))
                except: pass
            
            ws.column_dimensions[col_letter].width = max((max_len + 2), 12)
        wb.save(filename)
    except Exception as e:
        print(f"Error styling Excel: {e}")

def simpan_ke_excel_staff(list_transaksi, nama_staff, uang_tunai, uang_qris, total_setor, catatan):
    try:
        nama_file = get_nama_file_excel(nama_staff)
        
        data_row = {
            "TANGGAL": get_wib_now().strftime("%Y-%m-%d"),
            "JAM": get_wib_now().strftime("%H:%M"),
            "NAMA": nama_staff,
            "GEROBAK": list_transaksi[0]['GEROBAK'] if list_transaksi else "-",
            "CASH": uang_tunai,
            "QRIS": uang_qris,
            "TOTAL OMZET": total_setor,
        }

        for item in list_transaksi:
            if item['TIPE'] == 'JUAL':
                col_name = f"{item['KATEGORI']} - {item['ITEM']} ({int(item['HARGA'])})"
                data_row[col_name] = item['TERJUAL']

        data_row["CATATAN"] = catatan

        df_baru = pd.DataFrame([data_row])

        if os.path.exists(nama_file):
            df_lama = pd.read_excel(nama_file)
            df_final = pd.concat([df_lama, df_baru], ignore_index=True)
            df_final = df_final.fillna(0)
        else:
            df_final = df_baru.fillna(0)
            
        kolom_utama = ["TANGGAL", "JAM", "NAMA", "GEROBAK", "CASH", "QRIS", "TOTAL OMZET"]
        kolom_menu = [col for col in df_final.columns if col not in kolom_utama and col != "CATATAN"]
        
        urutan_final = kolom_utama + kolom_menu
        if "CATATAN" in df_final.columns:
            urutan_final.append("CATATAN")
            
        df_final = df_final[urutan_final]
        df_final.to_excel(nama_file, index=False)
            
        rapikan_excel(nama_file)
        return nama_file
    except Exception as e:
        st.error(f"âŒ Gagal Simpan Excel: {e}")
        return None

# ================= 5. TAMPILAN APLIKASI =================
def main():
    st.set_page_config(page_title="Sistem Gerobak Pro", page_icon="ğŸª", layout="centered")
    
    LOKASI_SEKARANG = get_lokasi_aktif()
    MENU_SEKARANG = get_menu_aktif()
    waktu_skrg = get_wib_now()
    
    st.title("OLEO DAIRY ğŸ®")
    st.caption(f"ğŸ“… {waktu_skrg.strftime('%d-%m-%Y %H:%M')} WIB")

    if 'user_nama' not in st.session_state: st.session_state['user_nama'] = None
    if 'user_pin' not in st.session_state: st.session_state['user_pin'] = None
    if 'keranjang_kasir' not in st.session_state: st.session_state['keranjang_kasir'] = {}

    # --- SIDEBAR LOGIN ---
    with st.sidebar:
        st.header("ğŸ” Akses Sistem")
        if st.session_state['user_nama'] is None:
            mode = st.radio("Pilih Menu:", ["Login Masuk", "Daftar Staff Baru"])
            if mode == "Login Masuk":
                pin = st.text_input("Masukkan PIN", type="password", max_chars=6)
                if st.button("Masuk Sistem"):
                    data_staff = load_json(FILE_DB_STAFF)
                    if pin == PIN_OWNER_LOGIN:
                        st.session_state['user_nama'] = "OWNER"; st.session_state['user_pin'] = PIN_OWNER_LOGIN; st.rerun()
                    elif pin in data_staff:
                        st.session_state['user_nama'] = data_staff[pin]; st.session_state['user_pin'] = pin; st.rerun()
                    else: st.error("PIN Tidak Dikenal")
            elif mode == "Daftar Staff Baru":
                nm = st.text_input("Nama Lengkap")
                pn = st.text_input("PIN (Angka)", max_chars=6)
                if st.button("Daftarkan Staff"): 
                    if not nm or not pn: st.error("Nama dan PIN wajib diisi!")
                    elif simpan_staff_baru(nm, pn): 
                        st.success(f"Staff {nm} Terdaftar!")
                    else: st.error("PIN Sudah Dipakai")
        else:
            st.success(f"Halo, {st.session_state['user_nama']}")
            if st.button("Keluar (Logout)"): 
                st.session_state['user_nama'] = None
                st.session_state['keranjang_kasir'] = {}
                st.rerun()

    # --- KONTEN UTAMA ---
    if st.session_state['user_nama']:
        user = st.session_state['user_nama']
        pin  = st.session_state['user_pin']

        # ================= FITUR OWNER =================
        if user == "OWNER":
            st.info("ğŸ”§ **MODE ADMIN / PEMILIK**")
            
            db_gerobak = load_json(FILE_DB_GEROBAK)
            ds = load_json(FILE_DB_STAFF)
            
            total_menu = sum(len(items) for items in MENU_SEKARANG.values())
            
            c1, c2, c3 = st.columns(3)
            c1.metric("Gerobak Buka", f"{len(db_gerobak)}")
            c2.metric("Total Staff", f"{len(ds)}")
            c3.metric("Total Menu", f"{total_menu}")
            
            t1, t2, t3, t4, t5, t6 = st.tabs(["ğŸ›’ Cek Toko", "ğŸ‘¥ Staff", "ğŸ“‹ Menu", "ğŸ“ Lokasi", "ğŸ“¥ Laporan", "ğŸšš Surat Jalan"])
            
            with t1: 
                if not db_gerobak: st.caption("Semua gerobak tutup.")
                for id_lok, nama_lok in LOKASI_SEKARANG.items():
                    info = db_gerobak.get(nama_lok)
                    status_icon = "ğŸŸ¢ TUTUP" if not info else "ğŸ”´ BUKA"
                    with st.expander(f"{status_icon} - {nama_lok}", expanded=bool(info)):
                        if info:
                            st.write(f"ğŸ‘¤ **Penjaga:** {info['pic']}")
                            st.write(f"â° **Buka:** {info['jam_masuk']}")
                            if st.button(f"â›” PAKSA TUTUP / RESET {nama_lok}", key=f"kick_{id_lok}"):
                                del db_gerobak[nama_lok]; save_json(FILE_DB_GEROBAK, db_gerobak); st.rerun()
            
            with t2: 
                st.dataframe(pd.DataFrame(list(ds.items()), columns=['PIN','NAMA']), hide_index=True, use_container_width=True)
                if ds:
                    pilih = st.selectbox("Hapus Staff:", [f"{v} ({k})" for k,v in ds.items()])
                    if st.button("Hapus"): hapus_staff(pilih.split('(')[1][:-1]); st.rerun()

            with t3: 
                st.write("**Kelola Menu Per Kategori**")
                c_kat, c_nama, c_harga = st.columns(3)
                kat_baru = c_kat.selectbox("Kategori:", ["Regular (Cup)", "Botol 250ml", "Botol 1 Liter"])
                nama_baru = c_nama.text_input("Nama Menu")
                harga_baru = c_harga.number_input("Harga", step=500)
                
                if st.button("ğŸ’¾ Simpan Menu"):
                    if nama_baru:
                        simpan_menu_baru(kat_baru, nama_baru, harga_baru)
                        st.success("Menu Disimpan!"); st.rerun()

                st.divider()
                for kat, items in MENU_SEKARANG.items():
                    with st.expander(f"ğŸ“‚ {kat} ({len(items)} menu)"):
                        if items:
                            df_menu = pd.DataFrame(list(items.items()), columns=['Nama Menu', 'Harga'])
                            st.dataframe(df_menu, hide_index=True, use_container_width=True)
                            to_del = st.selectbox(f"Hapus Menu di {kat}:", list(items.keys()), key=f"del_{kat}")
                            if st.button(f"Hapus {to_del}", key=f"btn_del_{kat}"):
                                hapus_menu(kat, to_del); st.rerun()

            with t4: 
                st.subheader("Daftar Gerobak")
                if LOKASI_SEKARANG:
                    df_lokasi = pd.DataFrame(list(LOKASI_SEKARANG.items()), columns=['ID', 'Nama Gerobak - Lokasi'])
                    st.dataframe(df_lokasi, hide_index=True, use_container_width=True)
                
                c_l1, c_l2 = st.columns([1,3])
                ids = [int(k) for k in LOKASI_SEKARANG.keys()]
                next_id = str(max(ids) + 1) if ids else "1"
                
                input_id = c_l1.text_input("ID", value=next_id)
                input_nama = c_l2.text_input("Nama Gerobak (Cth: Gerobak 01)")
                input_lokasi = st.text_input("ğŸ“ Lokasi (Cth: Unand)")
                
                if st.button("ğŸ’¾ Simpan Lokasi"):
                    if input_nama and input_lokasi:
                        nama_lengkap = f"{input_nama} - {input_lokasi}"
                        simpan_lokasi_baru(input_id, nama_lengkap)
                        st.rerun()

                if LOKASI_SEKARANG:
                    del_lok = st.selectbox("Hapus:", [f"{k} - {v}" for k,v in LOKASI_SEKARANG.items()])
                    if st.button("ğŸ—‘ï¸ Hapus"): hapus_lokasi(del_lok.split(' - ')[0]); st.rerun()

            with t5: 
                st.subheader("ğŸ“¥ Download Laporan")
                if ds:
                    list_nama_staff = list(ds.values())
                    pilih_staff_dl = st.selectbox("Pilih Staff:", list_nama_staff)
                    file_target = get_nama_file_excel(pilih_staff_dl)
                    
                    if os.path.exists(file_target):
                        with open(file_target, "rb") as file:
                            st.download_button(label=f"ğŸ“¥ Download {pilih_staff_dl}", data=file, file_name=file_target, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    else: st.warning("Belum ada laporan.")

                st.divider()
                st.error("âš ï¸ RESET SEMUA DATA (DANGER)")
                st.write("Menghapus: Staff, Menu, Lokasi, Surat Jalan, dan Semua Laporan.")
                
                with st.expander("ğŸ”´ Buka Menu Reset"):
                    password_reset = st.text_input("Masukkan Password Reset:", type="password")
                    
                    if st.button("ğŸ”¥ PIKIA-PIKIA BANA LUU"):
                        if password_reset == PASSWORD_RESET: 
                            try:
                                files_db = [FILE_DB_GEROBAK, FILE_DB_STAFF, FILE_DB_MENU, FILE_DB_LOKASI, FILE_DB_SURAT_JALAN]
                                for fdb in files_db:
                                    if os.path.exists(fdb): os.remove(fdb)
                                
                                file_excel = glob.glob("LAPORAN_*.xlsx")
                                for f in file_excel: os.remove(f)
                                
                                st.success("âœ… SYSTEM RESET BERHASIL!")
                                kirim_whatsapp("âš ï¸ *SYSTEM ALERT:* OWNER MELAKUKAN FULL RESET DATA.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Gagal Reset: {e}")
                        else:
                            st.error("â›” Password Salah!")
                            
            with t6:
                st.subheader("ğŸšš Buat Surat Jalan Baru")
                if not LOKASI_SEKARANG:
                    st.warning("Belum ada lokasi gerobak.")
                else:
                    sj_lokasi = st.selectbox("Tujuan Gerobak:", list(LOKASI_SEKARANG.values()), key="sj_lokasi")
                    
                    st.write("**Pilih Barang & Jumlah yang Dikirim:**")
                    sj_stok_input = {}
                    
                    for kategori, items in MENU_SEKARANG.items():
                        with st.expander(f"ğŸ“¦ Kirim {kategori}", expanded=False):
                            cols = st.columns(2)
                            for i, (m, hrg) in enumerate(items.items()):
                                key_input = f"{kategori}_{m}"
                                with cols[i%2]: 
                                    val = st.number_input(f"{m}", min_value=0, value=0, key=f"sj_kirim_{key_input}")
                                    if val > 0: sj_stok_input[key_input] = val 

                    if st.button("Kirim Surat Jalan"):
                        if not sj_stok_input:
                            st.error("Daftar barang tidak boleh kosong! Isi minimal 1 barang.")
                        else:
                            db_sj = load_json(FILE_DB_SURAT_JALAN)
                            id_sj = f"SJ-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                            
                            sj_catatan_text = ""
                            for k_item, jml in sj_stok_input.items():
                                kat_split, nama_split = k_item.split('_', 1)
                                sj_catatan_text += f"\nâ–«ï¸ [{kat_split}] {nama_split}: {jml}"

                            db_sj[id_sj] = {
                                "tanggal": get_wib_now().strftime("%Y-%m-%d %H:%M"),
                                "tujuan": sj_lokasi,
                                "barang_dict": sj_stok_input, 
                                "barang_text": sj_catatan_text, 
                                "status": "Menunggu Konfirmasi",
                                "penerima": "-"
                            }
                            save_json(FILE_DB_SURAT_JALAN, db_sj)
                            
                            tgl_sj = get_wib_now().strftime("%d-%m-%Y")
                            jam_sj = get_wib_now().strftime("%H:%M WIB")
                            kirim_whatsapp(f"ğŸšš *INFO SURAT JALAN BARU*\nğŸ“… Tanggal: {tgl_sj}\nâ° Waktu: {jam_sj}\nğŸ‘¤ Pengirim: {user} (Owner)\nğŸ“ Tujuan: {sj_lokasi}\n\n*Barang Dikirim:*{sj_catatan_text}\n\n_Menunggu konfirmasi staff._")
                            st.success("Surat Jalan berhasil dikirim ke staff!")
                            st.rerun()

                st.divider()
                st.write("**Riwayat Surat Jalan Terakhir**")
                db_sj = load_json(FILE_DB_SURAT_JALAN)
                if db_sj:
                    for id_sj, data_sj in reversed(list(db_sj.items())):
                        status_icon = "â³" if data_sj['status'] == "Menunggu Konfirmasi" else "âœ…"
                        with st.expander(f"{status_icon} {data_sj['tanggal']} | Tujuan: {data_sj['tujuan']}"):
                            st.write(f"**Barang:**{data_sj.get('barang_text', data_sj.get('barang', ''))}")
                            st.write(f"**Status:** {data_sj['status']} (Penerima']})")
                            if st.button(f"Hapus Riwayat", key=f"del_sj_{id_sj}"):
                                del db_sj[id_sj]
                                save_json(FILE_DB_SURAT_JALAN, db_sj)
                                st.rerun()
                else:
                    st.caption("Belum ada riwayat surat jalan.")

            st.divider()

        # ================= FITUR STAFF =================
        if not LOKASI_SEKARANG: st.error("Database Kosong"); st.stop()

        st.subheader("ğŸ“ Pilih Posisi Gerobak")
        pilihan_gerobak = st.selectbox("Pilih Gerobak & Lokasi:", list(LOKASI_SEKARANG.values()))
        
        db_gerobak = load_json(FILE_DB_GEROBAK)
        shift_aktif_di_lokasi = db_gerobak.get(pilihan_gerobak)

        is_lokasi_terisi = shift_aktif_di_lokasi is not None
        is_saya_di_sini  = is_lokasi_terisi and shift_aktif_di_lokasi['pin_pic'] == pin
        
        lokasi_lain_user = None
        for nama_g, data_g in db_gerobak.items():
            if data_g['pin_pic'] == pin and nama_g != pilihan_gerobak: lokasi_lain_user = nama_g; break

        if is_lokasi_terisi:
            if is_saya_di_sini: st.success(f"âœ… ANDA SEDANG AKTIF DI: {pilihan_gerobak}")
            else: st.error(f"â›” SEDANG DIPAKAI: {shift_aktif_di_lokasi['pic']}")
        else: st.info(f"ğŸŸ¢ {pilihan_gerobak} Kosong.")

        # --- CEK SURAT JALAN UNTUK STAFF ---
        db_sj = load_json(FILE_DB_SURAT_JALAN)
        ada_surat_jalan = False
        
        for id_sj, data_sj in db_sj.items():
            if data_sj['tujuan'] == pilihan_gerobak and data_sj['status'] == "Menunggu Konfirmasi":
                ada_surat_jalan = True
                st.warning("ğŸšš **ADA BARANG MASUK DARI GUDANG!**")
                st.info(f"**Daftar Tambahan Stok Jualan:**{data_sj.get('barang_text', data_sj.get('barang', ''))}")
                
                if st.button("âœ… Konfirmasi Terima Barang", key=f"terima_{id_sj}"):
                    data_sj['status'] = "Diterima"
                    data_sj['penerima'] = user
                    save_json(FILE_DB_SURAT_JALAN, db_sj)

                    pesan_tambahan_bot = ""
                    if is_saya_di_sini:
                        stok_aktif = db_gerobak[pilihan_gerobak].get('stok', {})
                        barang_masuk = data_sj.get('barang_dict', {})
                        
                        for k, v in barang_masuk.items():
                            if k in stok_aktif:
                                stok_aktif[k] += v  
                            else:
                                stok_aktif[k] = v   
                        
                        db_gerobak[pilihan_gerobak]['stok'] = stok_aktif
                        save_json(FILE_DB_GEROBAK, db_gerobak)
                        pesan_tambahan_bot = "\n\n_(Stok berhasil ditambahkan otomatis ke Kasir)_"

                    tgl_terima = get_wib_now().strftime("%d-%m-%Y")
                    jam_terima = get_wib_now().strftime("%H:%M WIB")
                    kirim_whatsapp(f"âœ… *LAPORAN TERIMA BARANG*\nğŸ“… Tanggal: {tgl_terima}\nâ° Waktu: {jam_terima}\nğŸ‘¤ Penerima: {user}\nğŸ“ Lokasi: {pilihan_gerobak}{pesan_tambahan_bot}")
                    st.success("Barang berhasil dikonfirmasi masuk!")
                    st.rerun()

        if ada_surat_jalan:
            st.error("âš ï¸ Anda wajib mengonfirmasi penerimaan Surat Jalan di atas sebelum bisa mengakses Kasir Gerobak.")
            st.stop() 

        # --- TABS SISTEM STAFF (DITAMBAH TAB KASIR) ---
        t_op, t_ks, t_cl = st.tabs(["â˜€ï¸ BUKA TOKO", "ğŸ’» KASIR (POS)", "ğŸŒ™ TUTUP TOKO"])

        with t_op:
            if lokasi_lain_user: st.error("âŒ Anda masih aktif di tempat lain.")
            elif is_lokasi_terisi and not is_saya_di_sini: st.error("ğŸ”’ Terkunci.")
            elif is_saya_di_sini: st.info("Toko sudah buka. Silakan pindah ke tab KASIR.")
            else:
                st.write("ğŸ“ **Persiapan Buka Toko (Input Stok)**")
                stok_input = {}
                
                for kategori, items in MENU_SEKARANG.items():
                    with st.expander(f"ğŸ“¦ Stok {kategori}", expanded=True):
                        cols = st.columns(2)
                        for i, (m, hrg) in enumerate(items.items()):
                            key_input = f"{kategori}_{m}"
                            with cols[i%2]: 
                                val = st.number_input(f"{m}", min_value=0, value=0, key=f"stok_{key_input}")
                                if val > 0: stok_input[key_input] = val 
                
                st.write("---")
                uang_kembalian_input = int(st.number_input("ğŸ’µ Uang Kembalian (Modal Kasir / Pecahan)", min_value=0, step=500, value=0, key="uang_kembalian_open"))
                
                if st.button("ğŸš€ BUKA SHIFT SEKARANG", key="btn_open"):
                    if not stok_input:
                        st.error("âš ï¸ Stok awal tidak boleh kosong! Wajib isi minimal 1 barang untuk buka toko.")
                    else:
                        tgl_skrg = get_wib_now().strftime("%d-%m-%Y")
                        jam_skrg_wib = get_wib_now().strftime("%H:%M WIB")
                        jam_masuk_db = get_wib_now().strftime("%H:%M")
                        
                        list_stok_text = ""
                        for k_item, jml in stok_input.items():
                            kat_split, nama_split = k_item.split('_', 1)
                            list_stok_text += f"\nğŸ“¦ [{kat_split}] {nama_split}: {jml}"

                        d = {
                            "tanggal": get_wib_now().strftime("%Y-%m-%d"), 
                            "jam_masuk": jam_masuk_db, 
                            "pic": user, 
                            "pin_pic": pin, 
                            "stok": stok_input,
                            "terjual": {},           # Untuk melacak penjualan kasir
                            "omzet_cash": 0,         # Untuk melacak cash real-time
                            "omzet_qris": 0,         # Untuk melacak qris real-time
                            "uang_kembalian": uang_kembalian_input
                        }
                        db_gerobak[pilihan_gerobak] = d; save_json(FILE_DB_GEROBAK, db_gerobak)
                        
                        kirim_whatsapp(f"â˜€ï¸ *LAPORAN OPENING SHIFT*\nğŸ“… Tanggal: {tgl_skrg}\nâ° Waktu: {jam_skrg_wib}\nğŸ‘¤ Nama: {user}\nğŸ“ Lokasi: {pilihan_gerobak}\nğŸ’µ *Uang Kembalian:* {format_rupiah(uang_kembalian_input)}\n\n*STOK AWAL:*{list_stok_text}")
                        st.success("Buka!"); st.rerun()

        # FITUR KASIR REAL-TIME
        with t_ks:
            if not is_saya_di_sini:
                st.warning("âš ï¸ Silakan BUKA TOKO terlebih dahulu.")
            else:
                c_menu, c_cart = st.columns([2, 1])
                
                with c_menu:
                    st.write("ğŸ›’ **Pilih Menu:**")
                    for kat, items in MENU_SEKARANG.items():
                        st.markdown(f"**{kat}**")
                        cols = st.columns(3)
                        for i, (m, hrg) in enumerate(items.items()):
                            key_unik = f"{kat}_{m}"
                            stok_tersedia = int(shift_aktif_di_lokasi['stok'].get(key_unik, 0))
                            
                            with cols[i%3]:
                                if st.button(f"{m}\n{format_rupiah(hrg)}\n(Stok: {stok_tersedia})", key=f"pos_{key_unik}", disabled=(stok_tersedia<=0), use_container_width=True):
                                    # Tambah ke keranjang
                                    if key_unik in st.session_state['keranjang_kasir']:
                                        if st.session_state['keranjang_kasir'][key_unik]['qty'] < stok_tersedia:
                                            st.session_state['keranjang_kasir'][key_unik]['qty'] += 1
                                        else: st.toast("âŒ Stok tidak cukup!")
                                    else:
                                        st.session_state['keranjang_kasir'][key_unik] = {'nama': m, 'kat': kat, 'harga': hrg, 'qty': 1}
                                    st.rerun()
                        st.divider()

                with c_cart:
                    st.write("ğŸ§¾ **Keranjang Belanja:**")
                    if not st.session_state['keranjang_kasir']:
                        st.info("Belum ada pesanan.")
                    else:
                        total_belanja = 0
                        for k, v in list(st.session_state['keranjang_kasir'].items()):
                            cc1, cc2 = st.columns([3, 1])
                            cc1.write(f"{v['nama']} (x{v['qty']})")
                            if cc2.button("âŒ", key=f"del_pos_{k}"):
                                del st.session_state['keranjang_kasir'][k]
                                st.rerun()
                            total_belanja += (v['harga'] * v['qty'])

                        st.markdown(f"### Total: {format_rupiah(total_belanja)}")
                        metode_bayar = st.radio("Metode Pembayaran:", ["Tunai (CASH)", "QRIS"])

                        if st.button("ğŸ’³ PROSES BAYAR", use_container_width=True, type="primary"):
                            # Tarik data dari DB untuk di-update
                            db_upd = load_json(FILE_DB_GEROBAK)
                            shift_upd = db_upd[pilihan_gerobak]

                            # Kurangi stok & Tambah data Terjual
                            for k_cart, v_cart in st.session_state['keranjang_kasir'].items():
                                qty_beli = v_cart['qty']
                                shift_upd['stok'][k_cart] -= qty_beli
                                shift_upd['terjual'][k_cart] = shift_upd.get('terjual', {}).get(k_cart, 0) + qty_beli

                            # Tambah ke Omzet Real-time
                            if "CASH" in metode_bayar:
                                shift_upd['omzet_cash'] = shift_upd.get('omzet_cash', 0) + total_belanja
                            else:
                                shift_upd['omzet_qris'] = shift_upd.get('omzet_qris', 0) + total_belanja

                            save_json(FILE_DB_GEROBAK, db_upd)
                            st.session_state['keranjang_kasir'] = {} # Kosongkan keranjang
                            st.toast("âœ… Transaksi Berhasil Dicatat!")
                            st.rerun()

        with t_cl:
            if not is_saya_di_sini:
                st.info("Toko belum dibuka atau bukan shift Anda.")
            else:
                st.write("ğŸ“ **Laporan Penjualan (Terekap Otomatis dari Kasir)**")
                
                uang_kembalian_awal = int(shift_aktif_di_lokasi.get('uang_kembalian', 0))
                data_terjual = shift_aktif_di_lokasi.get('terjual', {})
                omzet_cash_sistem = int(shift_aktif_di_lokasi.get('omzet_cash', 0))
                omzet_qris_sistem = int(shift_aktif_di_lokasi.get('omzet_qris', 0))
                omzet_total = omzet_cash_sistem + omzet_qris_sistem
                
                list_penjualan_temporary = [] 
                
                for kategori, items in MENU_SEKARANG.items():
                    with st.expander(f"ğŸ“¦ Status Barang {kategori}", expanded=False):
                        for m, harga_satuan in items.items():
                            key_unik = f"{kategori}_{m}"
                            
                            terjual_item = int(data_terjual.get(key_unik, 0))
                            sisa_stok_item = int(shift_aktif_di_lokasi['stok'].get(key_unik, 0))
                            
                            if terjual_item > 0 or sisa_stok_item > 0:
                                st.write(f"- {m} | Terjual: **{terjual_item}** | Sisa Stok: **{sisa_stok_item}**")

                            if terjual_item > 0:
                                list_penjualan_temporary.append({
                                    "KATEGORI": kategori, "ITEM": m, "HARGA": harga_satuan,
                                    "TERJUAL": terjual_item, "TIPE": "JUAL", "GEROBAK": pilihan_gerobak
                                })
                
                target_uang_fisik = omzet_cash_sistem + uang_kembalian_awal

                st.write("---")
                st.markdown(f"### ğŸ’° Total Omzet Penjualan (Murni): {format_rupiah(omzet_total)}")
                st.markdown(f"*(Sistem Mencatat Cash: {format_rupiah(omzet_cash_sistem)} | QRIS: {format_rupiah(omzet_qris_sistem)})*")
                
                if uang_kembalian_awal > 0:
                    st.info(f"ğŸ’µ Uang Kembalian (Modal Pagi): {format_rupiah(uang_kembalian_awal)}\n\n**ğŸ“Œ TARGET UANG FISIK DI LACI (Omzet Cash + Kembalian): {format_rupiah(target_uang_fisik)}**")
                else:
                    st.info(f"**ğŸ“Œ TARGET UANG FISIK DI LACI (Sama dengan Omzet Cash): {format_rupiah(target_uang_fisik)}**")
                
                c1, c2 = st.columns(2)
                uang_tunai = int(c1.number_input("Tunai Aktual (Hitung Total Uang Fisik di Laci)", step=500, value=target_uang_fisik, key="uang_tunai"))
                uang_qris = int(c2.number_input("QRIS Aktual (Cek mutasi/mesin EDC)", step=500, value=omzet_qris_sistem, key="uang_qris"))
                catatan = st.text_area("Catatan", key="catatan_closing")
                
                # Validasi Total Setor Aktual = Total Cash Aktual + Total Qris Aktual
                total_setor_aktual = uang_tunai + uang_qris
                
                # Uang yang seharusnya ada = (Cash Kasir + Modal) + Qris Kasir
                target_setor_keseluruhan = target_uang_fisik + omzet_qris_sistem
                selisih = total_setor_aktual - target_setor_keseluruhan
                
                if selisih == 0:
                    st.success("âœ… Uang Pas! Laporan Siap Dikirim.")
                    if st.button("ğŸ”’ TUTUP SHIFT & KIRIM", key="btn_close"):
                        
                        with st.spinner("Memproses..."):
                            
                            # Excel hanya menerima Omzet Murni (uang kembalian dikurangi dari cash laci)
                            excel_tunai = uang_tunai - uang_kembalian_awal
                            excel_total = excel_tunai + uang_qris
                            
                            # Tambahkan data dummy agar excel tidak error jika tidak ada penjualan sama sekali
                            if not list_penjualan_temporary:
                                list_penjualan_temporary.append({"KATEGORI": "-", "ITEM": "Tidak ada penjualan", "HARGA": 0, "TERJUAL": 0, "TIPE": "JUAL", "GEROBAK": pilihan_gerobak})

                            nama_file_excel = simpan_ke_excel_staff(
                                list_penjualan_temporary, user, excel_tunai, uang_qris, excel_total, catatan
                            )
                            
                            rincian_text = ""
                            for item in list_penjualan_temporary:
                                if item['TERJUAL'] > 0:
                                    rincian_text += f"\nâ–«ï¸ [{item['KATEGORI']}] {item['ITEM']}: {item['TERJUAL']}"
                            if not rincian_text: rincian_text = "\n(Nihil)"

                            tgl_tutup = get_wib_now().strftime("%d-%m-%Y")
                            jam_tutup = get_wib_now().strftime("%H:%M WIB")

                            # WA mengikuti format sebelumnya
                            msg = (f"ğŸŒ™ *LAPORAN CLOSING SHIFT*\n"
                                   f"ğŸ“… Tanggal: {tgl_tutup}\n"
                                   f"â° Waktu: {jam_tutup}\n"
                                   f"ğŸ‘¤ Nama: {user}\n"
                                   f"ğŸ“ Lokasi: {pilihan_gerobak}\n\n"
                                   f"ğŸ“Š *RINCIAN TERJUAL:*{rincian_text}\n\n"
                                   f"ğŸ’µ *Tunai Aktual:* {format_rupiah(uang_tunai)}\n"
                                   f"ğŸ’³ *QRIS Aktual:* {format_rupiah(uang_qris)}\n"
                                   f"ğŸ’° *Total Setor (Termasuk Modal):* {format_rupiah(total_setor_aktual)}\n"
                                   f"ğŸ“ *Catatan:* {catatan}")

                            kirim_whatsapp(msg)
                            
                            if pilihan_gerobak in db_gerobak:
                                del db_gerobak[pilihan_gerobak]; save_json(FILE_DB_GEROBAK, db_gerobak)
                        
                        st.balloons(); st.success("Selesai!"); st.rerun()
                else:
                    st.error(f"âš ï¸ **ADA SELISIH: {format_rupiah(selisih)}**")
                    st.warning("Tombol kirim terkunci. Pastikan uang Tunai & QRIS yang Anda input sama persis dengan Target Sistem.")

    else: st.info("â˜ï¸â˜ï¸ Login ada di Kiri atas ğŸ‘ˆğŸ‘ˆ")

if __name__ == "__main__":
    main()
